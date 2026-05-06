#!/usr/bin/env python3
"""
generate_excel.py
-----------------
Reads precomputed.json and pre-generates Excel files per gene × condition.

Output: data/excel/{GENE_ID}_{target}_{flags}.xlsx
      e.g.  data/excel/Os06g0275000_gene_110.xlsx
          data/excel/Os06g0275000_all_111.xlsx

Usage:
    python scripts/generate_excel.py
    python scripts/generate_excel.py --gene Os06g0275000
    python scripts/generate_excel.py --data-dir /data/holee/hap-browser/hap-browser/public/data
"""

import json
import os
import argparse
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Install with: pip install openpyxl")
    raise

# ── Utilities ─────────────────────────────────────────────────────────────────

REGION_LBL = {
    'cds': 'CDS', 'intron': 'Intron',
    'upstream': 'Upstream', 'downstream': 'Downstream',
    'utr5': "5'UTR", 'utr3': "3'UTR",
}

def classify_position(pos, gene_start, gene_end, cds_intervals):
    """Convert local pos → region string."""
    if pos < gene_start:
        return 'upstream'
    if pos > gene_end:
        return 'downstream'
    for (cs, ce) in cds_intervals:
        if cs <= pos <= ce:
            return 'cds'
    return 'intron'

def get_cds_intervals(features, gene_id):
    """Extract CDS intervals from GFF features (local coords)."""
    intervals = []
    for f in features:
        if f.get('type') == 'CDS':
            intervals.append((f['start'], f['end']))
    return intervals

def decode_enc(enc_str, alt_list, ref_base):
    """Convert enc char → allele."""
    if not enc_str:
        return ref_base
    c = enc_str
    if c == '0':
        return ref_base
    if c == '-':
        return '-'
    idx = int(c) - 1
    if idx < len(alt_list):
        return alt_list[idx]
    return ref_base

# ── Styles ────────────────────────────────────────────────────────────────────

BASE_COLORS = {
    'A': 'FF4ADE80',  # green
    'T': 'FFFBBF24',  # amber
    'G': 'FF818CF8',  # indigo
    'C': 'FFFB923C',  # orange
    '-': 'FFE2E8F0',  # gray (gap/del)
}
HEADER_FILL   = PatternFill('solid', fgColor='FF1E3A5F')
HEADER_FONT   = Font(color='FFFFFFFF', bold=True, size=9)
SUBHDR_FILL   = PatternFill('solid', fgColor='FFE2E8F0')
SUBHDR_FONT   = Font(bold=True, size=9)
THIN = Side(style='thin', color='FFD1D5DB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(cell, sub=False):
    cell.fill = SUBHDR_FILL if sub else HEADER_FILL
    cell.font = SUBHDR_FONT if sub else HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER

def style_base(cell, base):
    color = BASE_COLORS.get(base[:1] if base else '', 'FFFFFFFF')
    if base and base != ref_symbol:
        cell.fill = PatternFill('solid', fgColor=color)
        cell.font = Font(bold=True, size=9, color='FFFFFFFF')
    else:
        cell.font = Font(size=9)
    cell.alignment = Alignment(horizontal='center')
    cell.border = BORDER

ref_symbol = '__REF__'  # sentinel

# ── Main generation function ──────────────────────────────────────────────────

def generate_excel_for_gene(gene_id, data_dir, target='all', flags='111'):
    """Generate xlsx for a single condition."""
    pc_path = Path(data_dir) / 'precomputed' / f'{gene_id}.json'
    if not pc_path.exists():
        print(f"  precomputed not found: {pc_path}")
        return False

    with open(pc_path) as f:
        pc = json.load(f)

    samples    = pc['samples']
    offset     = pc['offset']
    region_len = pc['region_length']
    strand     = pc.get('strand', '+')
    seq        = pc.get('seq', '')
    gene_start_local = pc.get('gene_start', 0) - offset
    gene_end_local   = pc.get('gene_end', 0) - offset

    # positionData: enc schema
    pd_list = pc.get('positionData', [])
    pd_map  = {pd['pos']: pd for pd in pd_list}

    # determine combo key
    if target == 'all':
        # all = ignore gene range (whole region)
        combo_key = f'gene_{flags}'  # fallback to widest range
        # "all" handling: filter positions by flags directly
        snp_f  = flags[0] == '1'
        indel_f = flags[1] == '1'
        gap_f  = flags[2] == '1'
        positions = []
        for pd in pd_list:
            f = pd['f']
            match = False
            if snp_f  and (f & 1):  match = True
            if indel_f and (f & 10): match = True  # ins(8) + del(2)
            if gap_f  and (f & 4):  match = True
            if match:
                positions.append(pd['pos'])
    else:
        combo_key = f'{target}_{flags}'
        combo = pc.get('combos', {}).get(combo_key)
        if not combo:
            print(f"  combo not found: {combo_key}")
            return False
        positions = combo.get('variantPositions', [])
        haplotypes = combo.get('haplotypes', [])

    if not positions:
        print(f"  no positions for {gene_id} {target}_{flags}")
        return False

    # Rebuild haplotypes (for "all")
    if target == 'all':
        combo = pc.get('combos', {}).get(combo_key, {})
        haplotypes = combo.get('haplotypes', [])

    # sample → hap mapping
    sample_hap = {}
    for hap in haplotypes:
        for s in hap.get('samples', []):
            sample_hap[s] = hap['id']

    # ── Excel generation ──────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f'{gene_id[:12]}'

    gene_sym = pc.get('gene_sym', gene_id)

    # Row 1: gene info
    ws.cell(1, 1, 'Gene').font = Font(bold=True, size=9)
    ws.cell(1, 2, f'{gene_id} ({gene_sym}) {strand}  |  Region: {target.upper()}  |  Filters: '
            f'{"SNP " if snp_f else ""}{"InDel " if indel_f else ""}{"Gap" if gap_f else ""}')
    ws.cell(1, 2).font = Font(size=9)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=min(len(positions)+2, 20))

    # Row 2: header labels
    ROW_HAP  = 2
    ROW_POS  = 3
    ROW_REF  = 4
    ROW_ALTS = 5  # Alt sample
    ROW_ALTR = 6  # Alt read %
    DATA_ROW = 7

    for col_label, row_idx in [('Haplotype', ROW_HAP), ('RAP-DB position', ROW_POS),
                                ('Reference', ROW_REF), ('Alt sample', ROW_ALTS), ('Alt read', ROW_ALTR)]:
        cell = ws.cell(row_idx, 1, col_label)
        style_header(cell, sub=(row_idx > ROW_HAP))

    ws.cell(ROW_HAP, 2, 'Sample')
    style_header(ws.cell(ROW_HAP, 2), sub=False)
    ws.cell(ROW_POS, 2, 'Annotation')
    style_header(ws.cell(ROW_POS, 2), sub=True)
    for r in [ROW_REF, ROW_ALTS, ROW_ALTR]:
        ws.cell(r, 2, '')
        style_header(ws.cell(r, 2), sub=True)

    # position column headers
    for ci, pos in enumerate(positions):
        col = ci + 3
        rapdb_pos = pos + offset
        pd = pd_map.get(pos, {})
        f  = pd.get('f', 0)

        # Annotation
        region = classify_position(pos, gene_start_local, gene_end_local, [])
        if f & 32: region = 'cds'
        elif f & 16: region = 'intron'
        elif pos < gene_start_local: region = 'upstream'
        elif pos > gene_end_local: region = 'downstream'
        annot = REGION_LBL.get(region, region)

        ref_base = seq[pos - 1] if seq and 0 < pos <= len(seq) else 'N'

        # Row2: annotation
        c2 = ws.cell(ROW_HAP, col, annot)
        style_header(c2, sub=True)

        # Row 3: RAP-DB position (vertical)
        c3 = ws.cell(ROW_POS, col, rapdb_pos)
        c3.alignment = Alignment(horizontal='center', text_rotation=90)
        c3.font = Font(size=8)
        c3.border = BORDER

        # Row4: Reference
        c4 = ws.cell(ROW_REF, col, ref_base)
        style_header(c4, sub=True)

        # Row 5-6: alt info (blank if absent)
        ws.cell(ROW_ALTS, col, '').border = BORDER
        ws.cell(ROW_ALTR, col, '').border = BORDER

    # column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    for ci in range(len(positions)):
        ws.column_dimensions[get_column_letter(ci + 3)].width = 5

    # row heights
    ws.row_dimensions[ROW_POS].height = 65
    for r in [ROW_HAP, ROW_REF, ROW_ALTS, ROW_ALTR]:
        ws.row_dimensions[r].height = 14

    # sample data
    sample_idx_map = {s: i for i, s in enumerate(samples)}
    current_row = DATA_ROW
    prev_hap = None

    for hap in haplotypes:
        for sid in hap.get('samples', []):
            if sid not in sample_idx_map:
                continue
            si = sample_idx_map[sid]
            hap_id = hap['id']

            ws.cell(current_row, 1, hap_id if hap_id != prev_hap else '').font = Font(size=9)
            ws.cell(current_row, 2, sid).font = Font(size=9, family='Courier New')
            ws.cell(current_row, 1).border = BORDER
            ws.cell(current_row, 2).border = BORDER

            for ci, pos in enumerate(positions):
                col = ci + 3
                pd = pd_map.get(pos)
                ref_base = seq[pos - 1] if seq and 0 < pos <= len(seq) else 'N'
                if pd:
                    enc = pd.get('enc', '')
                    allele = decode_enc(enc[si] if si < len(enc) else '0',
                                       pd.get('alt', []), ref_base)
                else:
                    allele = ref_base

                cell = ws.cell(current_row, col, allele)
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(size=9, bold=(allele != ref_base and allele != '-'))
                if allele != ref_base:
                    base_color = BASE_COLORS.get(allele[:1], 'FFCCCCCC')
                    cell.fill = PatternFill('solid', fgColor=base_color)
                    cell.font = Font(size=9, bold=True, color='FFFFFFFF')
                cell.border = BORDER

            prev_hap = hap_id
            current_row += 1

    # output
    out_dir = Path(data_dir) / 'excel'
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f'{gene_id}_{target}_{flags}.xlsx'
    wb.save(out_path)
    print(f"  ✓ {out_path.name} ({current_row - DATA_ROW} samples, {len(positions)} positions)")
    return True


# ── CLI ───────────────────────────────────────────────────────────────────────

def find_data_dir():
    candidates = [
        Path(__file__).parent.parent / 'public' / 'data',
        Path('/data/holee/hap-browser/hap-browser/public/data'),
    ]
    for p in candidates:
        if (p / 'index.json').exists():
            return p
    return candidates[0]

def main():
    parser = argparse.ArgumentParser(description='Generate Excel files for HapBrowser')
    parser.add_argument('--data-dir', default=None)
    parser.add_argument('--gene', default=None, help='Single gene ID')
    parser.add_argument('--targets', default='all,gene,cds', help='Comma-separated targets')
    parser.add_argument('--flags', default='111,110,011,101,100,010,001',
                        help='Comma-separated flag combos (snp+indel+gap)')
    args = parser.parse_args()

    data_dir = Path(args.data_dir) if args.data_dir else find_data_dir()
    index_path = data_dir / 'index.json'
    if not index_path.exists():
        print(f"index.json not found in {data_dir}")
        return

    with open(index_path) as f:
        index = json.load(f)

    genes = []
    for group in index.get('groups', []):
        for g in group.get('genes', []):
            genes.append(g)

    if args.gene:
        genes = [g for g in genes if g['id'] == args.gene]

    targets = args.targets.split(',')
    flags_list = args.flags.split(',')

    print(f"Data dir: {data_dir}")
    print(f"Genes: {len(genes)}, Targets: {targets}, Flags: {flags_list}")
    print()

    total = 0
    for gene_info in genes:
        gene_id = gene_info['id']
        print(f"[{gene_id}]")
        for target in targets:
            for flags in flags_list:
                ok = generate_excel_for_gene(gene_id, data_dir, target, flags)
                if ok:
                    total += 1

    print(f"\nDone: {total} files generated → {data_dir}/excel/")

if __name__ == '__main__':
    main()
